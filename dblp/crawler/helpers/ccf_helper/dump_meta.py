#!/usr/bin/env python
# coding: utf-8
# 源 PDF 文件 http://www.ccf.org.cn/ccf/contentcore/resource/download?ID=32826
# 使用 https://smallpdf.com/pdf-to-excel 将 PDF 转换为 Excel
# 导出的 Meta Data 格式如下

# {
#    "<name>": {
#        "conference": {
#          "A" : [url1, url2, ],
#        }
#        "journal": {
#          "A" : [url1, url2, ],
#        }
#    }
# }

import json
import re
from openpyxl import load_workbook


def _remove_chars(target, del_chars=[u'（', u'）', ' ', u')', u'(']):
    s = target
    for c in del_chars:
        s = s.replace(c, '')
    return s


c_pattern = re.compile(r'.*([ABC]).*')

wb = load_workbook('CCF推荐国际学术会议和期刊目录（下载）.xlsx')
meta = {}
for name in wb.get_sheet_names():
    sheet = wb[name]
    # find the '序号' start line
    i = 0
    for idx in range(1, len(sheet['A']) + 1):
        if u'序号' in sheet['A{}'.format(idx)].value:
            i = idx
            break
    if i == 0:
        continue
    if i > 2:
        # create new category
        # format the name
        if u'期刊' in sheet['A{}'.format(i - 3)].value:
            s_type = "journal"
        elif u'会议' in sheet['A{}'.format(i - 3)].value:
            s_type = "conference"
        else:
            s_type = None
        category_name = _remove_chars(sheet['A{}'.format(i - 2)].value)
        category_name = category_name.replace(u'／', '/')
        # generate tags from name
        tags = category_name.split('/')

    classification = sheet['A{}'.format(i - 1)].value
    c_type = c_pattern.match(classification).group(1)
    di = {"s": 0, "f": 0, "p": 0, "u": 0}
    for row in sheet.iter_rows(min_row=i, max_row=i):
        for c in row:
            if c.value is None:
                continue
            elif u'简' in c.value:
                di["s"] = c.col_idx
            elif u'全' in c.value or u'名称' in c.value:  # fix
                di["f"] = c.col_idx
            elif u'出' in c.value:
                di["p"] = c.col_idx
            elif u'网' in c.value:
                di["u"] = c.col_idx
    # if not di['f']:
    #     for c in row:
    #         print c.value
    urls = []
    for n in range(i + 1, len(sheet['A']) + 1):
        short_name = sheet.cell(row=n, column=di['s']).value
        full_name = sheet.cell(row=n, column=di['f']).value
        publisher = sheet.cell(row=n, column=di['p']).value
        url = sheet.cell(row=n, column=di['u']).value
        if url is None:
            continue
        urls.append(url)
        if 'dblp' not in url:
            print short_name, ', ', full_name, ', ', sheet, s_type

    if category_name not in meta:
        meta[category_name] = {}
    if s_type not in meta[category_name]:
        meta[category_name][s_type] = {}
    if c_type not in meta[category_name][s_type]:
        meta[category_name][s_type][c_type] = []
    meta[category_name][s_type][c_type].extend(urls)

# final category
with open("meta.json", "w") as fout:
    fout.write(json.dumps(meta, ensure_ascii=False).encode('utf-8'))
